import json
import pytest
import openpyxl
import common.get_var as get_var

# df = openpyxl.load_workbook("/Users/bytedance/Downloads/testcase.xlsx")
# sheet = df["闭环转标注"]
# rows = sheet.max_row
# columns = sheet.max_column
#

test_data = []
def readxl(env, path):  # 提供环境和excel路径
    df = openpyxl.load_workbook(path)
    sheet = df["闭环转标注"]
    rows = sheet.max_row
    columns = sheet.max_column
    for x in range(3, rows + 1):
        list = []
        for y in range(4, 9):
            d = sheet.cell(x, y).value
            list.append(d)
        data = get_var.var(env, list[0], list[1])  # 根据excel里的环境，模块，账号，调用var方法，生成需要的host,header； 还需要method,url,data
        q = list[2:5]  # list是单行case，只截取method,url,data
        data.extend(q)  # 把生成的host，header，和截取的合并，最终生成测试数据：host,headers,method,url,data
        test_data.append(data)
    return test_data
    # print(len(test_data))


# headers = {
#     "Algorithm-Space":space_id,
#     "Content-Type":"application/json",
#     "token":token
# }
# headers["Algorithm-Space1"]=space_id
# headers["token"]=token
# print(headers)

if __name__ == "__main__":
    readxl("1", r"E:\pythonProjects\Learn\data\testcase.xlsx")
