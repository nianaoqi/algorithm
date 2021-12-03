import json
import pytest
import openpyxl
import yaml
import requests
import pymysql
import time
from hashlib import md5


# 发送请求
def send(host, method, url, data, headers):
    if method == "get":
        res = requests.get(host + url, params=data, headers=headers)
    elif method == "post":
        res = requests.post(host + url, json=data, headers=headers)
    elif method == "put":
        res = requests.put(host + url, json=data, headers=headers)
    elif method == "delete":
        res = requests.delete(host + url + "/" + data, headers=headers)
    return res


# 读excel，获取case
def readxl(env, path, sheet):  # 提供环境和excel路径
    df = openpyxl.load_workbook(path)
    sheet = df[sheet]
    rows = sheet.max_row
    columns = sheet.max_column
    test_data = []
    for x in range(3, rows + 1):
        l = []
        for y in range(4, 10):
            d = sheet.cell(x, y).value.strip()  # 获取单元格用例
            l.append(d)
        data = get_var(env, l[0], l[1])  # 获取host，header, biz_line
        data.extend(l[2:6])  # 加上method,url,data,assertion
        test_data.append(data)
    return test_data
    # print(test_data)


def writexl():
    a = 1


# 读yaml，返回json
def read_yaml(yaml_name, param):
    yml = open("./data/"+yaml_name, "r", encoding="utf8").read()
    yml = yaml.safe_load(yml)
    return yml[param]


def write_yaml(path, data):
    with open(path, "w", encoding="utf-8") as y:
        yaml.dump(data, y, allow_unicode=True)


# 读取环境变量，获取对应的host、space_id、menu_id、business_id，拼装需要的header和data
def get_var(env, module, username):  # 提供环境，模块，角色，生成对应header和space
    # 获取json格式的环境参数
    yml = open("/Users/bytedance/PycharmProjects/Learn/data/env_var.yaml", "r", encoding="utf-8").read()
    yml = yaml.safe_load(yml)

    host = yml[env]["host"]
    space_id = yml[env]["space_id"]
    biz_line = yml[env]["biz_line"]
    menu_id = yml[env][module]
    headers = {
        "Algorithm-Space": space_id,
        "Content-Type": "application/json"
    }
    data = {
        "username": None,
        "password": "123456",
        "code_id": "QSoYuUzE2s7f7cQDShVh",
        "code_value": "80"
    }

    # 生成token，boe、ppe8种，线上6种，共存在14种结果
    if env == 1 or 2:
        if username == "扫码用户":
            token = yml[env]["token"]
        else:
            data["username"] = username
            res = requests.post(host + "/arule/api/v1/user/login", json=data, headers=headers).json()
            token = res["data"]["token"]
    else:
        if username != "年奥齐":
            token = yml[env]["token"]
        else:
            token = yml[env]["token1"]

    # 生成ppe和非ppe的header
    if env == 2:
        headers = {
            "Algorithm-Space": space_id,
            "Content-Type": "application/json",
            "algorithm-menu": menu_id,
            "token": token,
            "X-TT-ENV": "ppe_evaluation",
            "x-use-ppe": "1"
        }
    else:
        headers = {
            "Algorithm-Space": space_id,
            "Content-Type": "application/json",
            "algorithm-menu": menu_id,
            "token": token
        }

    return [host, headers, biz_line]


def send(host, method, url, data, headers):
    if method == "get":
        res = requests.get(host + url, params=data, headers=headers)
    elif method == "post":
        res = requests.post(host + url, json=data, headers=headers)
    elif method == "put":
        res = requests.put(host + url, json=data, headers=headers)
    elif method == "delete":
        res = requests.delete(host + url + "/" + data, headers=headers)
    return res


def execute_sql(env, sql):
    yml = open("/Users/bytedance/PycharmProjects/Learn/data/database.yaml", "r", encoding="utf8").read()
    yml = yaml.safe_load(yml)
    db = pymysql.connect(
        host=yml[env]["host"],
        user=yml[env]["user"],
        password=yml[env]["password"],
        database=yml[env]["database"],
        # cursorclass=pymysql.cursors.DictCursor#指定为dict格式
    )
    cursor = db.cursor()
    sql = sql
    cursor.execute(sql)
    data = cursor.fetchall()
    return data


def get_md5(key):
    now_time = int(time.time())  # 获取秒级时间戳
    s = "key=" + str(key) + "&" + "time=" + str(now_time)  # 按"key=${app_key}&time=${time_unix}"组合
    sign = md5(s.encode("utf-8")).hexdigest()  # 转MD5，需先转为utf-8
    print(s)
    # print(sign)
    return sign
